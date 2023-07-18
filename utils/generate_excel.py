# -*- coding = utf-8 -*-
# @Time :2023/7/13 16:57
# @Author :小岳
# @Email  :401208941@qq.com
# @PROJECT_NAME :scenic_spots_comment
# @File :  generate_excel.py.py
import os

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rich.console import Console


def generate_excel(province: str, city: str, scene_name: str, comment: list, console: Console) -> None:
    """
    根据评论创建exile文件
    :return:
    """
    fail_count = 0
    script_path = os.path.abspath(__file__)
    grandparent_dir = os.path.dirname(os.path.dirname(script_path))
    path_file = os.path.join(grandparent_dir, f"data\{province}\{city}\{scene_name}")

    # 创建景区文件夹
    os.makedirs(path_file, exist_ok=True)
    path_excel = os.path.join(path_file, f"{scene_name}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.create_sheet(f"{scene_name}", index=0)
    ws.append(["id", "用户id", "昵称", "评论内容", "评论时间", "评分", "点赞数", "ip属地"])
    ws.row_dimensions[1].height = 24
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 120
    ws.column_dimensions['E'].width = 20
    for item in range(len(comment)):
        try:
            _id = comment[item]["commentId"]
            user_id = comment[item].get("userInfo", {}).get("userId", "")
            nick_name = comment[item].get("userInfo", {}).get("nickName", "")
            content = comment[item]["content"]
            comment_time = comment[item]["publishTypeTag"].replace(" 发布点评", "")
            score = comment[item].get("score", "")
            useful_vote_count = comment[item]["usefulCount"]
            user_ip = comment[item].get("ipLocatedName", "")
            ws.append([_id, user_id, nick_name, content, comment_time, score, useful_vote_count, user_ip])
        except Exception as e:
            fail_count += 1
    console.print(f"爬取评论结果：[red]失败数：{fail_count}[/red]，成功数：{len(comment) - fail_count}，共计：{len(comment)}", style="bold green")
    try:
        wb.save(path_excel)
        console.print(f"评论数据成功保存到 “{path_excel}”", style="bold green")
    except Exception as e:
        console.print(f"评论数据保存失败：{e}", style="bold red")

